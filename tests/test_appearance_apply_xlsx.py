# SPDX-License-Identifier: MIT
"""XLSX brand APPEARANCE apply tests (Cluster A, PR-3 / A2 - xlsx leg).

The xlsx peer of the docx ``ApplyTest`` / ``ApplySizeColorTest`` and the pptx
``PptxAppearanceApplyTest`` family: prove that the xlsx fill loops brand a filled
cell's font / size / color from the captured profile through the SHARED
``common.appearance`` orchestration and the xlsx ``XLSX_BACKEND``, under the same
invariants the docx/pptx legs hold:

  - **A cell IS the single run:** ``XLSX_BACKEND.runs_of`` yields ``[cell]`` and the
    per-axis writers reassign the (immutable) openpyxl font.
  - **Applied when unset:** a freshly-filled cell that inherits its formatting (the
    common case) is branded; the size writer inverts the capture-side
    ``round(sz * 2)`` via ``half_pts / 2`` points; a theme color token PREFERS the
    resolver-enriched concrete hex (not ``Color(theme=int)``).
  - **Set-only-when-unset:** a cell that already carries an explicit axis is never
    clobbered, so an inherited-but-correct workbook/style value is preserved.
  - **AFTER the cover-style re-assert:** appearance is applied only on a cell we wrote
    and after ``_reassert_cover_style``, so the named-style re-assert is not clobbered;
    a preserved-formula / merged slave cell is never touched.
  - **Empty appearance is a byte-identical no-op:** a pre-capture profile generates
    byte-for-byte what it generated before the apply path existed, and generate-twice
    on an appearance-carrying profile is byte-idempotent.

UNIVERSAL by construction: the synthetic workbooks are built in a temp dir (never
tuned to napoleon/ntt) and the branded values (Inter, 24 half-points, 1F4E79) are not
the openpyxl library defaults (Calibri / 11 / theme=1).
"""

from __future__ import annotations

import copy
import os
import sys
import tempfile
import unittest
from pathlib import Path

THIS = Path(__file__).resolve()
sys.path.insert(0, str(THIS.parent.parent / "scripts"))
sys.path.insert(0, str(THIS.parent))

from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import Color, Font  # noqa: E402
from openpyxl.workbook.defined_name import DefinedName  # noqa: E402

from brandkit.common import appearance  # noqa: E402
from brandkit.formats.xlsx import extract as xlsx_extract  # noqa: E402
from brandkit.formats.xlsx import generate as xg  # noqa: E402
from brandkit.grid.model import GridDocument  # noqa: E402
from brandkit.profile import schema, store  # noqa: E402
from brandkit.profile.resolver import ResolvedOp  # noqa: E402


# ---------------------------------------------------------------------------
# Synthetic shell + extract helpers (reuse the proven NumberFormatRoleTest shape)
# ---------------------------------------------------------------------------
def _build_shell(td: Path) -> Path:
    """A workbook with a single named cell ``title`` over ``Data!B1`` and a named
    region ``rows`` over ``Data!B3:B5`` - the brand guarantee gates every fill through
    these names. No literal brand font/color is in the shell (so a branded output is
    unambiguously the apply path's doing, not a passthrough)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["B1"] = "shell placeholder"
    wb.defined_names.add(DefinedName("title", attr_text="Data!$B$1"))
    wb.defined_names.add(DefinedName("rows", attr_text="Data!$B$3:$B$5"))
    shell = td / "shell.xlsx"
    wb.save(shell)
    return shell


def _extract(td: Path, shell: Path):
    old = os.getcwd()
    os.chdir(td)
    try:
        xlsx_extract.extract(shell, "syn", scope="project", cwd=td)
        return store.load_profile("syn", "project")
    finally:
        os.chdir(old)


def _with_body(
    profile: dict,
    *,
    latin: str | None = None,
    size_hp: int | None = None,
    color: dict | None = None,
) -> dict:
    """Inject captured document-body appearance into the SAME keys the capture engine
    fills (``theme.fonts.body`` / ``theme.text.body.color``) so the resolver carries the
    body default exactly as a really-captured profile would. Nothing is keyed off the
    workbook's identity, so the test stays universal."""
    body: dict = {}
    if latin:
        body["latin"] = latin
    if size_hp:
        body["size_hp"] = size_hp
    if body:
        profile["theme"].setdefault("fonts", {})["body"] = body
    if color:
        profile["theme"].setdefault("text", {})["body"] = {"color": color}
    return profile


# ---------------------------------------------------------------------------
# End-to-end apply through the fill loops
# ---------------------------------------------------------------------------
class XlsxAppearanceApplyTest(unittest.TestCase):
    def test_body_font_applied_to_filled_cell(self) -> None:
        """A freshly-filled named cell that inherits its formatting is branded with the
        captured body font (the body font default flows to every role; the resolver's
        family gate keeps the body size/color off a non-paragraph region role, so the
        font is the axis that lands here - parity in SHAPE, not richness)."""
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = _extract(td, _build_shell(td))
            profile = _with_body(
                loaded.profile,
                latin="Inter",
                size_hp=24,
                color={"kind": "hex", "hex": "1F4E79"},
            )
            grid = GridDocument(cells={"title": "Hello"})
            out = td / "out.xlsx"
            sink: list = []
            xg.generate(profile, loaded.shell_path, grid, out, findings=sink)
            cell = load_workbook(out)["Data"]["B1"]
            self.assertEqual(cell.value, "Hello")
            self.assertEqual(cell.font.name, "Inter")
            self.assertEqual(
                [f.check for f in sink if f.check == "appearance_color_skipped"], []
            )

    def test_body_font_applied_across_a_filled_region(self) -> None:
        """Every cell a region fill writes is branded; a cell the ragged grid did NOT
        write (a trailing region row) is left untouched (inherited)."""
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = _extract(td, _build_shell(td))
            profile = _with_body(loaded.profile, latin="Inter")
            grid = GridDocument(regions={"rows": [["a"], ["b"]]})  # B3, B4 (not B5)
            out = td / "out.xlsx"
            xg.generate(profile, loaded.shell_path, grid, out)
            ws = load_workbook(out)["Data"]
            self.assertEqual(ws["B3"].font.name, "Inter")
            self.assertEqual(ws["B4"].font.name, "Inter")
            # B5 was never written by the grid -> never branded (still inherited).
            self.assertFalse(ws["B5"].has_style)

    def test_empty_appearance_is_byte_identical_baseline(self) -> None:
        """A pre-capture profile (no body appearance) reads an empty op.appearance and
        is a no-op, so its bytes equal a second untouched generation (the regression
        net taken BEFORE relying on the apply path)."""
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = _extract(td, _build_shell(td))
            base = loaded.profile
            grid = GridDocument(cells={"title": "Hello"})
            a = td / "a.xlsx"
            b = td / "b.xlsx"
            xg.generate(copy.deepcopy(base), loaded.shell_path, grid, a)
            xg.generate(copy.deepcopy(base), loaded.shell_path, grid, b)
            self.assertEqual(a.read_bytes(), b.read_bytes())
            # And the filled cell carries NO direct style (pure inheritance) - proof the
            # no-op path never touched the font.
            self.assertFalse(load_workbook(a)["Data"]["B1"].has_style)

    def test_appearance_generation_is_byte_idempotent(self) -> None:
        """An appearance-carrying profile is byte-idempotent across two generations -
        the regression net for the new apply path."""
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = _extract(td, _build_shell(td))
            profile = _with_body(
                loaded.profile,
                latin="Inter",
                size_hp=24,
                color={"kind": "hex", "hex": "1F4E79"},
            )
            grid = GridDocument(
                cells={"title": "Hello"}, regions={"rows": [["a"], ["b"]]}
            )
            a = td / "a.xlsx"
            b = td / "b.xlsx"
            xg.generate(copy.deepcopy(profile), loaded.shell_path, grid, a)
            xg.generate(copy.deepcopy(profile), loaded.shell_path, grid, b)
            self.assertEqual(a.read_bytes(), b.read_bytes())


# ---------------------------------------------------------------------------
# XLSX_BACKEND probes + writers (the cell-is-a-run unit, all three axes)
# ---------------------------------------------------------------------------
def _op(appearance_dict: dict) -> ResolvedOp:
    return ResolvedOp(
        role_id="paragraph",
        resolver={},
        status="ok",
        confidence=1.0,
        kind="xlsx",
        appearance=appearance_dict,
    )


_FULL_APP = {
    "font": {"latin": "Inter"},
    "size_hp": 24,
    "color": {"kind": "hex", "hex": "111111"},
}


class XlsxBackendSetOnlyWhenUnsetTest(unittest.TestCase):
    """The ``XLSX_BACKEND`` ``*_unset`` probes drive set-only-when-unset directly,
    exercising the size unit conversion and color axes the family-gated end-to-end
    region role does not reach."""

    def _cell(self, value="x"):
        wb = Workbook()
        ws = wb.active
        c = ws.cell(row=1, column=1, value=value)
        return c

    def test_unset_axes_are_branded(self) -> None:
        cell = self._cell()  # unstyled: every axis inherited
        appearance.apply_role_appearance(xg.XLSX_BACKEND, cell, _op(_FULL_APP), [])
        self.assertEqual(cell.font.name, "Inter")
        # 24 half-points -> 12.0 pt; the round(sz*2) capture inverse holds.
        self.assertEqual(cell.font.sz, 12.0)
        self.assertEqual(cell.font.color.rgb, "FF111111")

    def test_size_unit_conversion_round_trips_half_points(self) -> None:
        cell = self._cell()
        appearance.apply_role_appearance(
            xg.XLSX_BACKEND, cell, _op({"size_hp": 36}), []
        )
        self.assertEqual(cell.font.sz, 18.0)  # 36 half-points, not 36 or 9

    def test_explicit_font_size_color_are_not_clobbered(self) -> None:
        cell = self._cell()
        cell.font = Font(name="Helvetica", size=14, color=Color(rgb="FFABCDEF"))
        appearance.apply_role_appearance(xg.XLSX_BACKEND, cell, _op(_FULL_APP), [])
        # None of the three explicit axes were overwritten.
        self.assertEqual(cell.font.name, "Helvetica")
        self.assertEqual(cell.font.sz, 14.0)
        self.assertEqual(cell.font.color.rgb, "FFABCDEF")

    def test_partial_explicit_font_brands_only_unset_axes(self) -> None:
        cell = self._cell()
        cell.font = Font(name="Helvetica")  # size + color genuinely unset
        appearance.apply_role_appearance(xg.XLSX_BACKEND, cell, _op(_FULL_APP), [])
        self.assertEqual(cell.font.name, "Helvetica")  # kept
        self.assertEqual(cell.font.sz, 12.0)  # branded
        self.assertEqual(cell.font.color.rgb, "FF111111")  # branded

    def test_empty_appearance_never_touches_the_cell(self) -> None:
        cell = self._cell()  # unstyled
        appearance.apply_role_appearance(xg.XLSX_BACKEND, cell, _op({}), [])
        self.assertFalse(cell.has_style)  # untouched: pure inheritance preserved


class XlsxSetCellColorTest(unittest.TestCase):
    """``_xlsx_set_cell_color`` writes hex as an opaque ``Color('FF'+hex)`` and PREFERS
    the resolver-enriched concrete hex for a theme token (not ``Color(theme=int)``,
    whose index order would be easy to misalign); a theme token with no resolvable hex,
    or a malformed hex, fails CLOSED (INFO, left inherited)."""

    def _cell(self):
        wb = Workbook()
        return wb.active.cell(row=1, column=1, value="x")

    def test_hex_ref_applies_as_opaque_argb(self) -> None:
        cell = self._cell()
        sink: list = []
        xg._xlsx_set_cell_color(cell, {"kind": "hex", "hex": "0B5394"}, sink)
        self.assertEqual(cell.font.color.rgb, "FF0B5394")
        self.assertEqual(cell.font.color.type, "rgb")
        self.assertEqual(sink, [])

    def test_theme_token_realized_via_enriched_hex(self) -> None:
        cell = self._cell()
        sink: list = []
        xg._xlsx_set_cell_color(
            cell, {"kind": "theme", "theme": "accent1", "hex": "2E7D32"}, sink
        )
        # Realized via the enriched hex, NOT Color(theme=int): an rgb-typed color.
        self.assertEqual(cell.font.color.rgb, "FF2E7D32")
        self.assertEqual(cell.font.color.type, "rgb")
        self.assertEqual(sink, [])

    def test_theme_token_without_hex_left_inherited_with_info(self) -> None:
        cell = self._cell()
        sink: list = []
        xg._xlsx_set_cell_color(cell, {"kind": "theme", "theme": "brandX"}, sink)
        self.assertFalse(cell.has_style)  # left inherited (never Color(theme=int))
        self.assertEqual([f.check for f in sink], ["appearance_color_skipped"])
        self.assertEqual(sink[0].severity, schema.Severity.INFO.value)

    def test_malformed_hex_fails_closed_with_info(self) -> None:
        cell = self._cell()
        sink: list = []
        xg._xlsx_set_cell_color(cell, {"kind": "hex", "hex": "zzzzzz"}, sink)
        self.assertFalse(cell.has_style)
        self.assertEqual([f.check for f in sink], ["appearance_color_skipped"])
        self.assertEqual(sink[0].severity, schema.Severity.INFO.value)


# ---------------------------------------------------------------------------
# Discipline: preserved-formula / merged slave cells are never branded
# ---------------------------------------------------------------------------
class XlsxAppearancePreservationTest(unittest.TestCase):
    def test_formula_cell_is_not_branded(self) -> None:
        """A named cell that already holds a shell formula is preserved verbatim:
        ``_fill_cell`` returns ``wrote=False`` so appearance is never applied and the
        formula cell keeps its own (inherited) formatting."""
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws["B1"] = "=1+1"  # a load-bearing shell formula
            wb.defined_names.add(DefinedName("title", attr_text="Data!$B$1"))
            shell = td / "shell.xlsx"
            wb.save(shell)
            loaded = _extract(td, shell)
            profile = _with_body(loaded.profile, latin="Inter")
            grid = GridDocument(cells={"title": "ignored"})
            out = td / "out.xlsx"
            xg.generate(profile, loaded.shell_path, grid, out)
            cell = load_workbook(out)["Data"]["B1"]
            self.assertEqual(cell.value, "=1+1")  # formula preserved
            self.assertNotEqual(cell.font.name, "Inter")  # not branded


if __name__ == "__main__":
    unittest.main()
