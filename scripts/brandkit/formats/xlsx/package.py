# SPDX-License-Identifier: MIT
"""Checked XLSX loading and preservation of unsupported worksheet extensions."""

from __future__ import annotations

import copy
import os
import posixpath
import tempfile
import warnings
import zipfile
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook as _openpyxl_load_workbook
from lxml import etree

from brandkit.ooxml import pack

_S_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
SPARKLINE_EXTENSION_URI = "{05C60535-1F16-4fd2-B633-F4F36F0B64E0}"
_SPARKLINE_WARNING = "Sparkline Group extension is not supported and will be removed"


def load_workbook_checked(path, **kwargs):
    """Validate ``path`` and load it while silencing one accounted-for warning.

    openpyxl cannot model x14 sparkline groups and warns whenever it sees them.
    Generation preserves those groups at the raw OOXML layer below, and QA compares
    the live shell/output extension inventory, so repeating that known warning at
    every read adds no signal.  Every other warning remains visible.
    """
    checked = pack.validate_package(path)
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=f"^{_SPARKLINE_WARNING}$",
            category=UserWarning,
            module=r"openpyxl\.worksheet\._reader",
        )
        return _openpyxl_load_workbook(checked, **kwargs)


def worksheet_extension_counts(path) -> dict[str, Counter[str]]:
    """Return ``sheet name -> Counter(extension URI)`` from live worksheet XML."""
    checked = pack.validate_package(path)
    result: dict[str, Counter[str]] = {}
    with zipfile.ZipFile(checked, "r") as zf:
        for sheet_name, part_name in _worksheet_parts(zf).items():
            root = pack.parse_xml_bytes(zf.read(part_name))
            ext_lst = root.find(f"{{{_S_NS}}}extLst")
            counts: Counter[str] = Counter()
            if ext_lst is not None:
                for ext in ext_lst.findall(f"{{{_S_NS}}}ext"):
                    uri = ext.get("uri")
                    if uri:
                        counts[uri] += 1
            result[sheet_name] = counts
    return result


def restore_supported_extensions(shell, output) -> list[str]:
    """Restore supported shell worksheet extensions lost by openpyxl.

    Currently only x14 sparkline groups are safe to splice: they are fully
    self-contained in the worksheet XML and reference ordinary cell ranges. Other
    extension families may depend on relationships/parts, so QA detects their loss
    and fails instead of copying an incomplete subtree.
    """
    shell_path = pack.validate_package(shell)
    output_path = pack.validate_package(output)
    shell_extensions: dict[str, list] = {}
    with zipfile.ZipFile(shell_path, "r") as zf:
        for sheet_name, part_name in _worksheet_parts(zf).items():
            root = pack.parse_xml_bytes(zf.read(part_name))
            ext_lst = root.find(f"{{{_S_NS}}}extLst")
            shell_extensions[sheet_name] = (
                []
                if ext_lst is None
                else [
                    copy.deepcopy(ext)
                    for ext in ext_lst.findall(f"{{{_S_NS}}}ext")
                    if ext.get("uri") == SPARKLINE_EXTENSION_URI
                ]
            )

    if not any(shell_extensions.values()):
        return []

    restored: list[str] = []
    fd, tmp_name = tempfile.mkstemp(
        prefix=output_path.name + ".", suffix=".tmp", dir=output_path.parent
    )
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        with (
            zipfile.ZipFile(output_path, "r") as zin,
            zipfile.ZipFile(tmp_path, "w") as zout,
        ):
            output_parts = _worksheet_parts(zin)
            part_to_sheet = {part: sheet for sheet, part in output_parts.items()}
            for info in zin.infolist():
                payload = zin.read(info.filename)
                output_sheet_name = part_to_sheet.get(info.filename)
                if output_sheet_name is not None and shell_extensions.get(
                    output_sheet_name
                ):
                    root = pack.parse_xml_bytes(payload)
                    ext_lst = root.find(f"{{{_S_NS}}}extLst")
                    if ext_lst is None:
                        ext_lst = etree.SubElement(root, f"{{{_S_NS}}}extLst")
                    present = Counter(
                        ext.get("uri")
                        for ext in ext_lst.findall(f"{{{_S_NS}}}ext")
                        if ext.get("uri")
                    )
                    for ext in shell_extensions[output_sheet_name]:
                        uri = ext.get("uri")
                        if present[uri] > 0:
                            present[uri] -= 1
                            continue
                        ext_lst.append(copy.deepcopy(ext))
                        restored.append(f"{output_sheet_name}:{uri}")
                    payload = etree.tostring(
                        root,
                        xml_declaration=True,
                        encoding="UTF-8",
                        standalone=True,
                    )
                zout.writestr(info, payload)
        os.replace(tmp_path, output_path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()
    return restored


def _worksheet_parts(zf: zipfile.ZipFile) -> dict[str, str]:
    workbook = pack.parse_xml_bytes(zf.read("xl/workbook.xml"))
    relationships = pack.parse_xml_bytes(zf.read("xl/_rels/workbook.xml.rels"))
    targets = {
        rel.get("Id"): rel.get("Target")
        for rel in relationships.findall(f"{{{_PKG_REL_NS}}}Relationship")
    }
    result: dict[str, str] = {}
    sheets = workbook.find(f"{{{_S_NS}}}sheets")
    if sheets is None:
        return result
    for sheet in sheets:
        name = sheet.get("name")
        rid = sheet.get(f"{{{_R_NS}}}id")
        target = targets.get(rid)
        if not name or not target:
            continue
        if target.startswith("/"):
            part = target.lstrip("/")
        else:
            part = posixpath.normpath(posixpath.join("xl", target))
        if part not in zf.namelist():
            raise pack.PackError(
                f"worksheet {name!r} relationship targets missing part {part!r}"
            )
        result[name] = part
    return result
